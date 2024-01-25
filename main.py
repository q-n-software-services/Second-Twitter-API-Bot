import time

import pandas
import requests

import openpyxl
import json

from deep_translator import GoogleTranslator

# To set your environment variables in your terminal run the following line:
# export 'BEARER_TOKEN'='<your_bearer_token>'
bearer_token = "AAAAAAAAAAAAAAAAAAAAAFpekAEAAAAANliUriJrNuBCotth2jby3fGiXlE%3DHMMwkgisYAjbl4pnibp6d7p9kDFSeWFohqf5F46lLv1vaOw2A9"

search_url = "https://api.twitter.com/2/spaces/search"


def create_url_tweet(id):
    tweet_fields = "tweet.fields=lang,author_id,possibly_sensitive,text,source,attachments,created_at"
    # Tweet fields are adjustable.
    # Options include:
    # attachments, author_id, context_annotations,
    # conversation_id, created_at, entities, geo, id,
    # in_reply_to_user_id, lang, non_public_metrics, organic_metrics,
    # possibly_sensitive, promoted_metrics, public_metrics, referenced_tweets,
    # source, text, and withheld
    ids = f"ids={id}"
    # You can adjust ids to include a single Tweets.
    # Or you can add to up to 100 comma-separated IDs
    url = "https://api.twitter.com/2/tweets?{}&{}".format(ids, tweet_fields)
    return url


def bearer_oauth(r):
    """
    Method required by bearer token authentication.
    """

    r.headers["Authorization"] = f"Bearer {bearer_token}"
    r.headers["User-Agent"] = "v2TweetLookupPython"
    return r


def connect_to_endpoint_tweet(url):
    response = requests.request("GET", url, auth=bearer_oauth)
    print(response.status_code)
    if response.status_code != 200:
        raise Exception(
            "Request returned an error: {} {}".format(
                response.status_code, response.text
            )
        )
    return response.json()


def create_headers(bearer_token):
    headers = {
        "Authorization": "Bearer {}".format(bearer_token),
        "User-Agent": "v2SpacesSearchPython"
    }
    return headers


def connect_to_endpoint(search_url, headers, params):
    response = requests.request("GET", search_url, headers=headers, params=params)
    print(response.status_code)
    if response.status_code != 200:
        raise Exception(response.status_code, response.text)
    return response.json()


def main():
    data_dict = {}
    file = openpyxl.load_workbook('keywords.xlsx')
    sheet = file.active
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            value = sheet.cell(row=i, column=j).value
            if value != None:

                print(value)

                search_term = value  # Replace this value with your search term
                # Optional params: created_at,creator_id,ended_at,host_ids,id,invited_user_ids,is_ticketed,lang,participant_count,scheduled_start,speaker_ids,started_at,state,subscriber_count,title,topic_ids,updated_at
                query_params = {'query': search_term, 'space.fields': 'title,id,host_ids,speaker_ids', 'expansions': 'creator_id'}
                headers = create_headers(bearer_token)
                json_response = connect_to_endpoint(search_url, headers, query_params)
                print(json.dumps(json_response, indent=4, sort_keys=True))
                print("\n\n\n")

                data_dict[value] = {value: json_response["meta"]["result_count"]}

                try:

                    translator = GoogleTranslator(source='es', target='en')
                    new_word = translator.translate(value)

                    print(f'{value}(es)\t -->\t {new_word}(en)')

                except:
                    new_word = value
                    print(f"Phrase : \t {value}\t NOT Translated so is used/searched as it is.")

                search_term = new_word # Replace this value with your search term
                # Optional params: host_ids,conversation_controls,created_at,creator_id,id,invited_user_ids,is_ticketed,lang,media_key,participants,scheduled_start,speaker_ids,started_at,state,title,updated_at
                query_params = {'query': search_term, 'space.fields': 'title,created_at', 'expansions': 'creator_id'}
                headers = create_headers(bearer_token)
                json_response = connect_to_endpoint(search_url, headers, query_params)
                print(json.dumps(json_response, indent=4, sort_keys=True))
                print("\n\n\n\n\n")

                data_dict[value][search_term] = json_response["meta"]["result_count"]

    print(data_dict)

    data_list = []

    for key in data_dict:
        data_list.append(data_dict[key])

    print(data_list)

    print("list = ", len(data_list))
    print("Dictionary = ", len(data_dict))

    csv_file = open("data.csv", 'w')
    csv_file.write("Spanish Phrase, Tweet Count (es), English Translation, Tweet Count (en)\n")
    plot_data = []
    for term in data_list:
        my_line = ''
        name = ''
        count = 0
        for n, key in enumerate(term):
            my_line += f"{key}, {term[key]}, "
            if n == 0:
                name += key
                name += ' (es) / '
            if n == 1:
                name += ' (en) '
            count += int(term[key])

            # csv_file.write(f"{term.keys()}, {term[term.keys()]}, {term.keys()[1]}, {term[term.keys()[1]]}\n")
            # csv_file.write(f"{term.keys()}, {term[term.keys()[0]]}, Same word , 0\n")
        plot_data.append({"Keyword": name, "Count": count})
        my_line += '\n'
        csv_file.write(my_line)

    csv_file.close()

    fig_df = pandas.DataFrame.from_dict(plot_data)

    # import plotly.express as px
    #
    # fig = px.bar(fig_df, x='Keyword', y='Count')
    # fig.show()


url = create_url_tweet(1619914907583135747)
json_response = connect_to_endpoint_tweet(url)
print(json.dumps(json_response, indent=4, sort_keys=True))

while True:
    main()
    time.sleep(1200)
